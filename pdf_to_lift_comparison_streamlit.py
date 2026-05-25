"""
PDF to Lift Comparison Excel Generator
======================================
Streamlit app to upload a consultant specification PDF and multiple vendor offer PDFs,
extract key elevator comparison data, and generate a professional Excel comparison file.

How to run:
1) Save this file as app.py
2) Install packages:
   pip install streamlit pymupdf openpyxl pandas
3) Run:
   streamlit run app.py

Notes:
- This is a rule-based extraction engine designed for elevator/vendor offers like KONE, TKE, EEE, AG MELCO.
- It works best with text-based PDFs. Scanned PDFs need OCR before upload.
- After extraction, the app lets you review/edit the data before exporting to Excel.
"""

from __future__ import annotations

import io
import re
from dataclasses import dataclass, asdict
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import fitz  # PyMuPDF
import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


# ------------------------------------------------------------
# Data models
# ------------------------------------------------------------

@dataclass
class LiftItem:
    vendor: str = ""
    source_file: str = ""
    tower: str = ""
    lift_group: str = ""
    model: str = ""
    origin: str = ""
    qty: Optional[int] = None
    capacity_kg: Optional[int] = None
    persons: Optional[int] = None
    speed_ms: Optional[float] = None
    floors: Optional[int] = None
    stops: Optional[int] = None
    openings: Optional[int] = None
    travel_height_m: Optional[float] = None
    car_size: str = ""
    door_size: str = ""
    shaft_size: str = ""
    pit_depth_mm: Optional[int] = None
    overhead_mm: Optional[int] = None
    code: str = ""
    drive_type: str = ""
    machine_room: str = ""
    doors: str = ""
    fire_rated_door: str = ""
    finishing: str = ""
    compliance_status: str = "Clarification Required"
    remarks: str = ""
    unit_price_aed: Optional[float] = None
    total_price_aed: Optional[float] = None


@dataclass
class VendorCommercial:
    vendor: str = ""
    source_file: str = ""
    project_name: str = ""
    offer_ref: str = ""
    offer_date: str = ""
    validity: str = ""
    warranty: str = ""
    free_maintenance: str = ""
    delivery: str = ""
    installation: str = ""
    payment_terms: str = ""
    total_price_aed: Optional[float] = None
    additional_items_aed: Optional[float] = None
    vat_note: str = ""
    exclusions: str = ""
    deviations: str = ""
    recommendation_note: str = ""


# ------------------------------------------------------------
# PDF extraction
# ------------------------------------------------------------

@st.cache_data(show_spinner=False)
def extract_pdf_text(file_bytes: bytes) -> Tuple[str, List[str]]:
    """Return full text and page-wise text from a PDF."""
    doc = fitz.open(stream=file_bytes, filetype="pdf")
    pages = []
    for page in doc:
        text = page.get_text("text") or ""
        pages.append(text)
    full_text = "\n".join(pages)
    return normalize_text(full_text), [normalize_text(p) for p in pages]


def normalize_text(text: str) -> str:
    text = text.replace("\u2013", "-").replace("\u2014", "-")
    text = text.replace("\u00a0", " ")
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


# ------------------------------------------------------------
# Helpers
# ------------------------------------------------------------

def first_match(patterns: List[str], text: str, flags=re.I | re.S) -> str:
    for pat in patterns:
        m = re.search(pat, text, flags)
        if m:
            return clean_value(m.group(1))
    return ""


def clean_value(v: str) -> str:
    return re.sub(r"\s+", " ", str(v)).strip(" :-\n\t")


def to_float(value: str) -> Optional[float]:
    if value is None:
        return None
    s = str(value).replace(",", "")
    m = re.search(r"-?\d+(?:\.\d+)?", s)
    return float(m.group(0)) if m else None


def to_int(value: str) -> Optional[int]:
    f = to_float(value)
    return int(round(f)) if f is not None else None


def money_to_float(value: str) -> Optional[float]:
    if not value:
        return None
    s = value.replace(",", "")
    nums = re.findall(r"\d+(?:\.\d+)?", s)
    if not nums:
        return None
    return float(nums[-1])


def detect_vendor(filename: str, text: str) -> str:
    hay = f"{filename}\n{text[:3000]}".lower()
    if "kone" in hay:
        return "KONE"
    if "tke" in hay or "tk elevator" in hay or "thyssenkrupp" in hay:
        return "TKE"
    if "eee" in hay or "elevator engineering enterprises" in hay:
        return "EEE"
    if "ag melco" in hay or "agmelco" in hay or "melco elevator" in hay:
        return "AG MELCO"
    if "otis" in hay:
        return "OTIS"
    if "schindler" in hay:
        return "SCHINDLER"
    if "mitsubishi" in hay:
        return "MITSUBISHI"
    return Path(filename).stem[:25].upper()


def extract_project_name(text: str) -> str:
    return first_match([
        r"Project Name\s*:?\s*(.+?)(?:\n|Attention|Client|$)",
        r"Project\s*:?\s*(Radiant Bridges.+?)(?:\n|Consultant|Sub:|$)",
        r"(RADIANT BRIDGES TOWERS.+?)(?:\n|ADM|CLIENT|$)",
    ], text)


# ------------------------------------------------------------
# Specification extraction
# ------------------------------------------------------------

def extract_spec_reference(filename: str, text: str) -> Dict[str, str]:
    """Extract main consultant/spec reference requirements as a dict."""
    spec = {
        "source_file": filename,
        "project_name": extract_project_name(text),
        "code": first_match([r"EN\s*81[- ]20", r"(EN\s*81[- ]20)"], text) or "EN81-20 / EN81-50 as applicable",
        "maintenance": first_match([r"maintenance.*?period of\s*(\d+\s*months)", r"free maintenance\s*:?\s*(\d+\s*months)"], text),
        "fire_lift_note": "Fireman lift to comply with local Civil Defense / UAE Fire Code and EN81-72 where applicable.",
    }

    # Build general required groups from common Radiant Bridges spec patterns.
    groups = []
    for tower, lift_group, cap, speed in [
        ("Tower A", "PL1~PL2", 1350, 4.0),
        ("Tower A", "PL3~PL5", 1600, 4.0),
        ("Tower A", "FL1", 1600, 4.0),
        ("Tower B", "PL1~PL2", 1350, 4.0),
        ("Tower B", "PL3~PL5", 1600, 4.0),
        ("Tower B", "FL1", 1600, 4.0),
        ("Tower C", "PL1~PL2", 1350, 3.5),
        ("Tower C", "PL3", 1350, 3.5),
        ("Tower C", "FL1", 1600, 4.0),
    ]:
        groups.append({
            "tower": tower,
            "lift_group": lift_group,
            "capacity_kg": cap,
            "speed_ms": speed,
            "code": spec["code"],
            "required_status": "Consultant / Specification Requirement",
        })

    spec["groups"] = groups
    return spec


# ------------------------------------------------------------
# Vendor technical extraction
# ------------------------------------------------------------

def extract_commercial(filename: str, vendor: str, text: str) -> VendorCommercial:
    c = VendorCommercial(vendor=vendor, source_file=filename)
    c.project_name = extract_project_name(text)
    c.offer_ref = first_match([
        r"(?:Ref|Reference|Our reference|Quote Reference|Offer)\s*:?\s*([A-Z0-9/_.-]+)",
        r"Tender number\s*:?\s*([A-Z0-9/_.-]+)",
    ], text)
    c.offer_date = first_match([
        r"Date\s*:?\s*(\d{1,2}[./-]\d{1,2}[./-]\d{2,4})",
        r"(\d{1,2}\s+[A-Za-z]+\s+\d{4})",
        r"(\d{1,2}[./-]\d{1,2}[./-]\d{4})",
    ], text)
    c.validity = first_match([
        r"valid(?:ity)?(?: for)?\s*(?:a period of)?\s*([^.\n]*?\d+\s*(?:days|months)[^.\n]*)",
        r"Price validity\s*:?\s*([^.\n]+)",
    ], text)
    c.warranty = first_match([r"Warranty\s*:?\s*([^.\n]*?\d+\s*months)", r"warranty.*?(\d+\s*months)"], text)
    c.free_maintenance = first_match([r"Free\s+maintenance\s*:?\s*([^.\n]*?\d+\s*months)", r"maintenance.*?(\d+\s*months)"], text)
    c.delivery = first_match([
        r"Material Delivery.*?(within\s*\d+\s*months[^.\n]*)",
        r"Delivery Program.*?(\d+\s*weeks[^.\n]*)",
        r"delivery.*?(\d+\s*months[^.\n]*)",
    ], text)
    c.installation = first_match([
        r"Installation\s*:?\s*(will be done within\s*\d+\s*months[^.\n]*)",
        r"installation.*?(\d+\s*months[^.\n]*)",
    ], text)

    # Payment terms: capture a compact block around payment terms.
    pay = first_match([
        r"Payment Terms?.{0,80}((?:\n|.){0,900}?)(?:Warranty|Validity|Terms and Conditions|Exclusions|$)",
        r"payment.{0,80}((?:\n|.){0,700}?)(?:Warranty|Validity|$)",
    ], text)
    c.payment_terms = clean_value(pay)[:1200]

    # Total values: find AED amounts, choose largest as likely offer total.
    amounts = [money_to_float(x) for x in re.findall(r"(?:AED|Dhs\.?|Dirhams)?\s*\d[\d,]*\.\d{2}\s*(?:AED|Dhs\.?|Dirhams)?", text, re.I)]
    amounts = [a for a in amounts if a and a > 10000]
    if amounts:
        c.total_price_aed = max(amounts)

    # Deviation / notes block.
    c.deviations = first_match([
        r"Notes and deviations from tender specification\s*((?:\n|.){0,1500}?)(?:Preliminary delivery|Scope of works|Pricing|$)",
        r"Notes\s*:?\s*((?:\n|.){0,1200}?)(?:Yours faithfully|Table of Contents|$)",
    ], text)[:1500]

    c.exclusions = first_match([
        r"Exclusions?\s*:?\s*((?:\n|.){0,1200}?)(?:Payment|Warranty|Validity|$)",
        r"Not included\s*:?\s*((?:\n|.){0,800}?)(?:Payment|Warranty|$)",
    ], text)[:1200]

    return c


def split_into_lift_blocks(text: str) -> List[str]:
    """Try to split vendor text into blocks containing lift technical tables/items."""
    patterns = [
        r"(?:Elevator specifications|Elevator Specifications|Equipment No\.).{0,5000}",
        r"(?:Tower\s+[ABC].{0,80}(?:PL|FL).{0,3500})",
        r"(?:Item\s+\d+\s+Tower\s+[ABC].{0,1200})",
    ]
    blocks = []
    for pat in patterns:
        blocks.extend(re.findall(pat, text, re.I | re.S))
    # Add commercial table lines as separate chunks.
    for line in text.splitlines():
        if re.search(r"Tower\s+[ABC].*(PL|FL|Fire|Passenger)", line, re.I):
            blocks.append(line)
    # Deduplicate roughly.
    seen = set()
    clean_blocks = []
    for b in blocks:
        b = clean_value(b)
        key = b[:200]
        if len(b) > 20 and key not in seen:
            seen.add(key)
            clean_blocks.append(b)
    return clean_blocks


def extract_items(filename: str, vendor: str, text: str) -> List[LiftItem]:
    items: List[LiftItem] = []

    # 1) Extract commercial line items, which usually have clean group/qty/price.
    commercial_patterns = [
        r"(Tower\s+[ABC]\s*\([^\)]*(?:PL|FL)[^\)]*\).*?)(?=\n\s*\d+\s+Tower|\n\s*\d+\s+FL|\n\s*\d{1,3},\d|$)",
        r"(FL1\s*\(Tower\s+[ABC]\).*?)(?=\n\s*\d+\s+Tower|\n\s*\d+\s+FL|\n\s*\d{1,3},\d|$)",
    ]
    chunks = []
    for pat in commercial_patterns:
        chunks.extend(re.findall(pat, text, re.I | re.S))

    for ch in chunks:
        item = parse_lift_chunk(ch, vendor, filename)
        if item and item.lift_group:
            items.append(item)

    # 2) Extract technical grouped tables by known tower/group patterns.
    known_groups = [
        ("Tower A", "PL1~PL2"), ("Tower A", "PL3~PL5"), ("Tower A", "FL1"),
        ("Tower B", "PL1~PL2"), ("Tower B", "PL3~PL5"), ("Tower B", "FL1"),
        ("Tower C", "PL1~PL2"), ("Tower C", "PL3"), ("Tower C", "FL1"),
        ("Tower C", "PL1~PL3"),
    ]

    for tower, group in known_groups:
        block = find_near_group_block(text, tower, group)
        if block:
            tech_item = parse_lift_chunk(block, vendor, filename)
            tech_item.tower = tower
            tech_item.lift_group = group
            # Merge with commercial item if same group exists.
            merged = False
            for existing in items:
                if same_group(existing, tech_item):
                    merge_items(existing, tech_item)
                    merged = True
                    break
            if not merged:
                items.append(tech_item)

    # 3) Vendor-specific fallback for KONE/TKE/EEE/AG MELCO.
    if not items:
        for block in split_into_lift_blocks(text):
            item = parse_lift_chunk(block, vendor, filename)
            if item and (item.tower or item.lift_group or item.capacity_kg or item.speed_ms):
                items.append(item)

    # Postprocess compliance and dedupe.
    unique: List[LiftItem] = []
    for item in items:
        enrich_item_from_text(item, text)
        item.compliance_status = determine_compliance(item, text)
        item.remarks = build_item_remarks(item, text)
        if not any(same_group(x, item) and x.vendor == item.vendor for x in unique):
            unique.append(item)
        else:
            for x in unique:
                if same_group(x, item) and x.vendor == item.vendor:
                    merge_items(x, item)
                    break
    return unique


def find_near_group_block(text: str, tower: str, group: str, radius: int = 1800) -> str:
    escaped_tower = re.escape(tower).replace("\\ ", r"\s+")
    group_variants = [group, group.replace("~", "-"), group.replace("~", " to "), group.replace("~", "")]
    for gv in group_variants:
        pat = escaped_tower + r".{0,80}" + re.escape(gv).replace("\\~", r"\s*[~-]\s*")
        m = re.search(pat, text, re.I | re.S)
        if m:
            start = max(0, m.start() - 500)
            end = min(len(text), m.end() + radius)
            return text[start:end]
    return ""


def parse_lift_chunk(ch: str, vendor: str, filename: str) -> Optional[LiftItem]:
    ch = normalize_text(ch)
    item = LiftItem(vendor=vendor, source_file=filename)

    # Tower/group.
    tower_match = re.search(r"Tower\s*([ABC])", ch, re.I)
    if tower_match:
        item.tower = f"Tower {tower_match.group(1).upper()}"

    group_match = re.search(r"(PL\s*\d\s*(?:[~\-]\s*PL?\d)?|PL\s*\d\s*[~\-]\s*PL?\d|FL\s*\d|Fire\s*man\s*Lift|Fireman\s*Lift)", ch, re.I)
    if group_match:
        group = group_match.group(1).upper().replace(" ", "")
        group = group.replace("-", "~")
        group = re.sub(r"PL(\d)~(\d)", r"PL\1~PL\2", group)
        if "FIRE" in group:
            group = "FL1"
        item.lift_group = group

    # Model/origin/type.
    item.model = first_match([
        r"Model\s*:?\s*([A-Za-z0-9\-_/ ]{2,40})",
        r"Type\s*:?\s*([A-Za-z0-9\-_/ ]{2,50})",
        r"\b(LEHY-M-II|MRA-QS6417|zeta200\s*MR|MiniSpace|MonoSpace)\b",
    ], ch)
    item.origin = first_match([r"Country of Origin\s*:?\s*([A-Za-z ]+)", r"Supply source\s*-?\s*([A-Za-z ()]+)"], ch)
    item.code = first_match([r"Code/?\s*Regulation compliance\s*:?\s*([A-Z0-9\- ]+)", r"(EN\s*81[- ]20[- ]?50|EN\s*81[- ]20)", r"Applicable code\s*:?\s*([A-Z0-9/\- ]+)",], ch)

    # Main numbers.
    item.qty = to_int(first_match([r"Number of Units\s*:?\s*(\d+)", r"QTY\s*(\d+)", r"\s(\d+)\s+[\d,]+\.\d{2}\s*AED"], ch))
    item.capacity_kg = to_int(first_match([r"(?:Load|Capacity)\s*\(?kg\)?\s*:?\s*(\d+)", r"\b(1350|1600|1050|1275|2000)\b\s*(?:Kg|kg)?"], ch))
    item.persons = to_int(first_match([r"(?:Persons|Passengers number)\s*:?\s*(\d+)", r"(18|21)\s*Persons"], ch))
    item.speed_ms = to_float(first_match([r"Speed\s*\(?m/s\)?\s*:?\s*(\d+(?:\.\d+)?)", r"\b(3\.5|4\.0|4)\s*(?:m/s|mps)\b"], ch))
    item.floors = to_int(first_match([r"Number of Floors\s*:?\s*(\d+)", r"Floors\s+(\d+)"], ch))
    item.stops = to_int(first_match([r"Number of Stops\s*:?\s*(\d+)", r"Stops\s+(\d+)"], ch))
    item.openings = to_int(first_match([r"(?:Openings|No\. of Openings)\s*:?\s*(\d+)"], ch))
    item.travel_height_m = to_float(first_match([r"Travel Height\s*\(?m\)?\s*:?\s*(\d+(?:\.\d+)?)", r"Total Travel\s*:?\s*([\d,]+)\s*mm"], ch))
    if item.travel_height_m and item.travel_height_m > 1000:
        item.travel_height_m = round(item.travel_height_m / 1000, 3)

    # Sizes.
    item.car_size = first_match([
        r"Car (?:inner )?size\s*W\s*x\s*D\s*x\s*H\s*:?\s*([\d,]+\s*mm.*?[\d,]+\s*mm.*?[\d,]+\s*mm)",
        r"Car Width.*?(\d{3,4}).*?Car Depth.*?(\d{3,4}).*?Car Total Height.*?(\d{3,4})",
    ], ch)
    if re.fullmatch(r"\d+", item.car_size or ""):
        pass
    item.door_size = first_match([
        r"Doors? opening size\s*W\s*x\s*H\s*:?\s*([\d,]+\s*mm.*?[\d,]+)",
        r"Door Width.*?(\d{3,4}).*?Door Height.*?(\d{3,4})",
    ], ch)
    item.shaft_size = first_match([
        r"(?:Hoistway|Shaft) (?:inner )?size\s*W\s*x\s*D\s*:?\s*([\d,]+\s*mm.*?[\d,]+\s*mm)",
        r"Shaft Width.*?(\d{3,4}).*?Shaft Depth.*?(\d{3,4})",
    ], ch)
    item.pit_depth_mm = to_int(first_match([r"Pit Depth\s*\(?mm\)?\s*:?\s*(\d+)", r"Pit depth\s*:?\s*(\d+)\s*mm"], ch))
    item.overhead_mm = to_int(first_match([r"Overhead Height\s*\(?mm\)?\s*:?\s*(\d+)", r"Overhead\s*:?\s*(\d+)\s*mm"], ch))

    # Door/fire/finish.
    item.doors = first_match([r"Landing door - Panel Type\s*:?\s*([^\n]+)", r"Doors opening mode\s*:?\s*([^\n]+)", r"Door opening type\s*:?\s*([^\n]+)"], ch)
    item.fire_rated_door = first_match([r"Landing door - Fire Rated\s*:?\s*([^\n]+)", r"(2\s*HRS\s*FIRE\s*RATED\s*DOOR[^\n]*)", r"(E120)"], ch)
    item.finishing = first_match([r"Landing Door Decoration[^:]*:?\s*([^\n]+)", r"Cabin.*?(Hairline Stainless Steel[^\n]+)", r"Car door\s*([^\n]+)"], ch)

    # Prices.
    item.unit_price_aed = money_to_float(first_match([r"(\d[\d,]*\.\d{2})\s*AED\s+\d[\d,]*\.\d{2}\s*AED", r"UNIT\s+TOTAL.*?(\d[\d,]*\.\d{2})"], ch))
    money_values = [money_to_float(x) for x in re.findall(r"\d[\d,]*\.\d{2}\s*AED", ch, re.I)]
    money_values = [x for x in money_values if x]
    if money_values:
        item.total_price_aed = max(money_values)

    if not any([item.tower, item.lift_group, item.capacity_kg, item.speed_ms, item.model, item.total_price_aed]):
        return None
    return item


def same_group(a: LiftItem, b: LiftItem) -> bool:
    return bool(a.tower and b.tower and a.lift_group and b.lift_group and a.tower == b.tower and a.lift_group == b.lift_group)


def merge_items(base: LiftItem, new: LiftItem) -> None:
    for field, value in asdict(new).items():
        if field in ["vendor", "source_file"]:
            continue
        if getattr(base, field) in [None, "", 0] and value not in [None, "", 0]:
            setattr(base, field, value)


def enrich_item_from_text(item: LiftItem, full_text: str) -> None:
    if not item.machine_room:
        if re.search(r"machine room", full_text, re.I):
            item.machine_room = "Machine Room / MR mentioned"
        elif re.search(r"MRL|machine room less", full_text, re.I):
            item.machine_room = "MRL mentioned"
    if not item.drive_type:
        item.drive_type = first_match([r"Drive type\s*:?\s*([^\n]+)", r"(gearless machine[^\n]*)", r"(VVVF[^\n]*)"], full_text)


def determine_compliance(item: LiftItem, full_text: str) -> str:
    flags = []
    if item.speed_ms is not None:
        if item.lift_group.startswith("PL") and item.tower in ["Tower A", "Tower B"] and item.speed_ms < 4.0:
            flags.append("Speed below A/B requirement")
        if item.lift_group.startswith("FL") and item.speed_ms < 4.0:
            flags.append("Fire lift speed below requirement")
    if item.lift_group.startswith("FL"):
        if re.search(r"not compliant|not comply|not.*local fire|not.*civil defense", full_text, re.I):
            flags.append("Fire lift compliance disclaimer found")
        if item.car_size and re.search(r"1700.*2000|1400.*2200", item.car_size):
            flags.append("Fire lift car size to be checked")
    if re.search(r"deviation|clarification|required|not included|by others", full_text[:10000], re.I):
        flags.append("Deviations/By others noted")
    if flags:
        return "Clarification Required"
    return "Generally Complied - Subject to Final Technical Review"


def build_item_remarks(item: LiftItem, full_text: str) -> str:
    remarks = []
    if item.lift_group.startswith("FL"):
        remarks.append("Verify UAE Fire Code/Civil Defense stretcher and adjacent-door requirements.")
    if item.pit_depth_mm:
        remarks.append(f"Pit depth offered: {item.pit_depth_mm} mm.")
    if item.overhead_mm:
        remarks.append(f"Overhead offered: {item.overhead_mm} mm.")
    if re.search(r"ID by Others|by customer|bare finish|decoration.*by others", full_text, re.I):
        remarks.append("Cabin/door ID or flooring may be by others; confirm scope.")
    if re.search(r"separation beam|separator beam", full_text, re.I):
        remarks.append("Separator beam item mentioned; confirm inclusion/additional cost.")
    if item.vendor == "KONE" and re.search(r"not compliant.*Fire", full_text, re.I):
        remarks.append("KONE note/disclaimer found regarding Tower A/B fire lift cabin compliance.")
    return " ".join(dict.fromkeys(remarks))


# ------------------------------------------------------------
# Comparison dataframe generation
# ------------------------------------------------------------

def build_comparison_df(spec: Dict[str, str], vendor_items: List[LiftItem]) -> pd.DataFrame:
    vendors = sorted(set(i.vendor for i in vendor_items))
    rows = []
    parameters = [
        ("Capacity", "capacity_kg"),
        ("Speed", "speed_ms"),
        ("Quantity", "qty"),
        ("Floors", "floors"),
        ("Stops", "stops"),
        ("Travel Height", "travel_height_m"),
        ("Car Size", "car_size"),
        ("Door Size", "door_size"),
        ("Shaft Size", "shaft_size"),
        ("Pit Depth", "pit_depth_mm"),
        ("Overhead", "overhead_mm"),
        ("Model", "model"),
        ("Country of Origin", "origin"),
        ("Code / Standard", "code"),
        ("Machine Room", "machine_room"),
        ("Door / Fire Rating", "fire_rated_door"),
        ("Finishing", "finishing"),
        ("Compliance Status", "compliance_status"),
        ("Remarks", "remarks"),
        ("Unit Price AED", "unit_price_aed"),
        ("Total Price AED", "total_price_aed"),
    ]

    groups = spec.get("groups", [])
    if not groups:
        # fallback groups from extracted items
        groups = sorted({(i.tower, i.lift_group) for i in vendor_items if i.tower and i.lift_group})
        groups = [{"tower": t, "lift_group": g, "capacity_kg": "", "speed_ms": "", "code": ""} for t, g in groups]

    for g in groups:
        tower = g.get("tower", "")
        lift_group = g.get("lift_group", "")
        for label, attr in parameters:
            row = {
                "Tower": tower,
                "Lift Group": lift_group,
                "Parameter": label,
                "Consultant / Spec": spec_value_for_parameter(g, label, attr),
            }
            for v in vendors:
                item = find_item(vendor_items, v, tower, lift_group)
                val = getattr(item, attr, "") if item else ""
                row[v] = format_cell_value(val, attr)
            rows.append(row)
    return pd.DataFrame(rows)


def spec_value_for_parameter(group: Dict[str, str], label: str, attr: str) -> str:
    if attr == "capacity_kg":
        return f"{group.get('capacity_kg', '')} kg" if group.get("capacity_kg") else ""
    if attr == "speed_ms":
        return f"{group.get('speed_ms', '')} m/s" if group.get("speed_ms") else ""
    if attr == "code":
        return str(group.get("code", ""))
    if label == "Compliance Status":
        return "Required to comply"
    return "As per drawings/specification"


def find_item(items: List[LiftItem], vendor: str, tower: str, lift_group: str) -> Optional[LiftItem]:
    # exact
    for i in items:
        if i.vendor == vendor and i.tower == tower and i.lift_group == lift_group:
            return i
    # group fallback for PL1~PL3 vs PL1~PL2/PL3
    for i in items:
        if i.vendor == vendor and i.tower == tower:
            if lift_group in i.lift_group or i.lift_group in lift_group:
                return i
            if lift_group.startswith("PL") and i.lift_group.startswith("PL") and tower == i.tower:
                pass
    return None


def format_cell_value(val, attr: str) -> str:
    if val is None:
        return ""
    if isinstance(val, float):
        if attr in ["speed_ms", "travel_height_m"]:
            return f"{val:g}"
        if "price" in attr:
            return f"{val:,.2f}"
    if isinstance(val, int):
        if attr in ["capacity_kg"]:
            return f"{val} kg"
        if attr in ["pit_depth_mm", "overhead_mm"]:
            return f"{val} mm"
        return str(val)
    return str(val)


def build_commercial_df(commercials: List[VendorCommercial]) -> pd.DataFrame:
    rows = []
    for c in commercials:
        rows.append({
            "Vendor": c.vendor,
            "Project Name": c.project_name,
            "Offer Ref": c.offer_ref,
            "Offer Date": c.offer_date,
            "Validity": c.validity,
            "Warranty": c.warranty,
            "Free Maintenance": c.free_maintenance,
            "Delivery": c.delivery,
            "Installation": c.installation,
            "Payment Terms": c.payment_terms,
            "Total Price AED": c.total_price_aed,
            "Additional Items AED": c.additional_items_aed,
            "VAT / Notes": c.vat_note,
            "Exclusions": c.exclusions,
            "Deviations / Notes": c.deviations,
            "Source File": c.source_file,
        })
    return pd.DataFrame(rows)


def build_summary_df(commercials: List[VendorCommercial], items: List[LiftItem]) -> pd.DataFrame:
    rows = []
    for c in commercials:
        vendor_items = [i for i in items if i.vendor == c.vendor]
        clarification_count = sum(1 for i in vendor_items if "Clarification" in i.compliance_status)
        total_qty = sum(i.qty or 0 for i in vendor_items)
        rows.append({
            "Vendor": c.vendor,
            "Total Offered Price AED": c.total_price_aed,
            "Detected Lift Qty": total_qty if total_qty else "To verify",
            "Warranty / Free Maintenance": c.warranty or c.free_maintenance,
            "Delivery": c.delivery,
            "Installation": c.installation,
            "Technical Status": "Clarification Required" if clarification_count else "Generally Complied - Subject to Review",
            "Key Risk / Comment": vendor_summary_comment(c.vendor, c, vendor_items),
        })
    df = pd.DataFrame(rows)
    if not df.empty and "Total Offered Price AED" in df.columns:
        df = df.sort_values(by="Total Offered Price AED", na_position="last")
    return df


def vendor_summary_comment(vendor: str, commercial: VendorCommercial, items: List[LiftItem]) -> str:
    text = " ".join([commercial.deviations or "", commercial.exclusions or "", " ".join(i.remarks for i in items)])
    comments = []
    if re.search(r"fire.*not compliant|not compliant.*fire|Civil Defense", text, re.I):
        comments.append("Fireman lift compliance to be resolved before award.")
    if re.search(r"by others|ID by Others|flooring", text, re.I):
        comments.append("Finishing/flooring scope by others to be confirmed.")
    if re.search(r"separation|separator", text, re.I):
        comments.append("Separator beam cost/scope to be confirmed.")
    if not comments:
        comments.append("Review final technical compliance, exclusions, payment terms, and delivery before award.")
    return " ".join(comments)


# ------------------------------------------------------------
# Excel writer
# ------------------------------------------------------------

def write_excel(
    output: io.BytesIO,
    project_name: str,
    summary_df: pd.DataFrame,
    comparison_df: pd.DataFrame,
    commercial_df: pd.DataFrame,
    items_df: pd.DataFrame,
    source_notes_df: pd.DataFrame,
) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    write_df_sheet(wb, "Executive Summary", summary_df, title=f"{project_name or 'Project'} - Lift Vendor Comparison")
    write_df_sheet(wb, "Technical Comparison", comparison_df, title="Technical Comparison Matrix")
    write_df_sheet(wb, "Commercial Comparison", commercial_df, title="Commercial Comparison")
    write_df_sheet(wb, "Extracted Raw Data", items_df, title="Extracted Raw Lift Data")
    write_df_sheet(wb, "Source Notes", source_notes_df, title="Source Notes / Extraction Log")

    add_recommendation_block(wb["Executive Summary"], summary_df)
    apply_workbook_style(wb)
    wb.save(output)
    output.seek(0)


def write_df_sheet(wb: Workbook, sheet_name: str, df: pd.DataFrame, title: str = ""):
    ws = wb.create_sheet(sheet_name)
    if title:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(1, len(df.columns)))
        cell = ws.cell(row=1, column=1, value=title)
        cell.font = Font(bold=True, size=14, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 25
        start_row = 3
    else:
        start_row = 1

    # headers
    for col_idx, col in enumerate(df.columns, 1):
        c = ws.cell(row=start_row, column=col_idx, value=col)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="5B9BD5")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # data
    for r_idx, row in enumerate(df.itertuples(index=False), start_row + 1):
        for c_idx, value in enumerate(row, 1):
            if pd.isna(value):
                value = ""
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if isinstance(value, (int, float)) and "AED" in str(ws.cell(start_row, c_idx).value).upper():
                cell.number_format = '#,##0.00'

    # freeze/filter
    ws.freeze_panes = ws.cell(row=start_row + 1, column=1).coordinate
    if len(df.columns) > 0 and len(df) > 0:
        ws.auto_filter.ref = f"A{start_row}:{get_column_letter(len(df.columns))}{start_row + len(df)}"

    # widths
    for col_idx, col in enumerate(df.columns, 1):
        width = min(max(len(str(col)) + 2, 12), 42)
        if any(k in str(col).lower() for k in ["remarks", "notes", "payment", "deviation", "exclusion"]):
            width = 38
        if str(col).lower() in ["parameter", "consultant / spec"]:
            width = 24
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def add_recommendation_block(ws, summary_df: pd.DataFrame) -> None:
    start = ws.max_row + 3
    ws.merge_cells(start_row=start, start_column=1, end_row=start, end_column=8)
    title = ws.cell(start, 1, "Engineer Recommendation / Review Notes")
    title.font = Font(bold=True, color="FFFFFF")
    title.fill = PatternFill("solid", fgColor="7030A0")
    title.alignment = Alignment(horizontal="center")

    notes = [
        "1. Final award should be subject to full compliance with consultant specification, project drawings, UAE Fire Code/Civil Defense requirements, and EN81 standards.",
        "2. Fireman lift cabin size, door arrangement, stretcher compliance, pit depth, overhead, and shaft compatibility must be formally confirmed by the selected vendor.",
        "3. Confirm all exclusions, by-others items, separator beam scope/cost, ID finishes, flooring, landing jamb/door finishes, warranty, free maintenance, and payment terms before final negotiation.",
        "4. Lowest commercial offer should not be selected unless technical compliance and local authority requirements are fully satisfied.",
    ]
    for idx, note in enumerate(notes, start + 1):
        ws.merge_cells(start_row=idx, start_column=1, end_row=idx, end_column=8)
        c = ws.cell(idx, 1, note)
        c.alignment = Alignment(wrap_text=True, vertical="top")

    # Add status dropdown for summary technical status.
    dv = DataValidation(type="list", formula1='"Complied,Clarification Required,Deviation,Not Offered"', allow_blank=True)
    ws.add_data_validation(dv)
    for row in range(4, max(4, ws.max_row + 1)):
        # try to apply to column 7 if that is Technical Status
        if ws.cell(3, 7).value == "Technical Status":
            dv.add(ws.cell(row, 7))


def apply_workbook_style(wb: Workbook) -> None:
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                cell.border = border
                if cell.row > 3:
                    cell.font = Font(size=10)
        # row heights
        for row_idx in range(1, ws.max_row + 1):
            ws.row_dimensions[row_idx].height = 22
        # conditional fills
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    v = cell.value.lower()
                    if "clarification required" in v or "not compliant" in v or "deviation" in v:
                        cell.fill = PatternFill("solid", fgColor="FFF2CC")
                    elif "complied" in v or "generally complied" in v:
                        cell.fill = PatternFill("solid", fgColor="E2F0D9")
                    elif "not offered" in v:
                        cell.fill = PatternFill("solid", fgColor="F4CCCC")
        ws.sheet_view.showGridLines = False


# ------------------------------------------------------------
# Streamlit UI
# ------------------------------------------------------------

def main():
    st.set_page_config(page_title="Lift PDF Comparison Generator", layout="wide")
    st.title("Lift PDF Comparison Generator")
    st.caption("Upload consultant specification and vendor offer PDFs. Review extracted data, then export a professional Excel comparison.")

    with st.sidebar:
        st.header("Upload PDFs")
        spec_file = st.file_uploader("Consultant Specification PDF", type=["pdf"], key="spec")
        vendor_files = st.file_uploader("Vendor Offer PDFs", type=["pdf"], accept_multiple_files=True, key="vendors")
        st.divider()
        project_override = st.text_input("Project name override", value="Radiant Bridges Towers")
        generate = st.button("Generate Comparison", type="primary")

    if not generate:
        st.info("Upload the specification and vendor PDFs, then click Generate Comparison.")
        return

    if not vendor_files:
        st.error("Please upload at least one vendor offer PDF.")
        return

    source_notes = []

    spec = {"project_name": project_override, "groups": []}
    if spec_file:
        with st.spinner("Reading specification PDF..."):
            spec_text, spec_pages = extract_pdf_text(spec_file.getvalue())
            spec = extract_spec_reference(spec_file.name, spec_text)
            if project_override:
                spec["project_name"] = project_override
            source_notes.append({
                "File": spec_file.name,
                "Type": "Specification",
                "Detected Vendor": "Consultant / Specification",
                "Pages": len(spec_pages),
                "Extracted Characters": len(spec_text),
                "Notes": "Specification requirements extracted.",
            })
    else:
        # Default groups for Radiant Bridges if spec not uploaded.
        spec = extract_spec_reference("No specification uploaded", project_override)
        spec["project_name"] = project_override

    all_items: List[LiftItem] = []
    commercials: List[VendorCommercial] = []

    with st.spinner("Extracting vendor offers..."):
        for vf in vendor_files:
            text, pages = extract_pdf_text(vf.getvalue())
            vendor = detect_vendor(vf.name, text)
            commercial = extract_commercial(vf.name, vendor, text)
            if project_override:
                commercial.project_name = project_override
            items = extract_items(vf.name, vendor, text)
            commercials.append(commercial)
            all_items.extend(items)
            source_notes.append({
                "File": vf.name,
                "Type": "Vendor Offer",
                "Detected Vendor": vendor,
                "Pages": len(pages),
                "Extracted Characters": len(text),
                "Notes": f"Extracted {len(items)} lift item records. Review data before final use.",
            })

    comparison_df = build_comparison_df(spec, all_items)
    commercial_df = build_commercial_df(commercials)
    summary_df = build_summary_df(commercials, all_items)
    items_df = pd.DataFrame([asdict(i) for i in all_items])
    source_notes_df = pd.DataFrame(source_notes)

    st.success("Extraction completed. Please review below before downloading Excel.")

    tab1, tab2, tab3, tab4 = st.tabs(["Executive Summary", "Technical Comparison", "Commercial", "Raw Extracted Data"])
    with tab1:
        edited_summary = st.data_editor(summary_df, use_container_width=True, num_rows="dynamic")
    with tab2:
        edited_comparison = st.data_editor(comparison_df, use_container_width=True, num_rows="dynamic", height=520)
    with tab3:
        edited_commercial = st.data_editor(commercial_df, use_container_width=True, num_rows="dynamic", height=420)
    with tab4:
        edited_items = st.data_editor(items_df, use_container_width=True, num_rows="dynamic", height=420)
        st.dataframe(source_notes_df, use_container_width=True)

    output = io.BytesIO()
    write_excel(
        output=output,
        project_name=spec.get("project_name", project_override),
        summary_df=edited_summary,
        comparison_df=edited_comparison,
        commercial_df=edited_commercial,
        items_df=edited_items,
        source_notes_df=source_notes_df,
    )

    filename = f"Lift_Comparison_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    st.download_button(
        label="Download Excel Comparison",
        data=output,
        file_name=filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
    )


if __name__ == "__main__":
    main()
